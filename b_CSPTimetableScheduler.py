import pandas as pd
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import random

class CSPTimetableScheduler:
    """
    Main scheduler class that handles loading data, constraint checking,
    and timetable generation using CSP backtracking algorithm.
    """


    """
        functions
        1. constructor✅
        2. get_consecutive_slot_pairs✅
        3. find_suitable_rooms✅
        4. find_suitable_labs✅
        5. is_slot_available✅
        6. is_actual_lab✅
        7. is_course_already_scheduled_today✅
        8. get_theory_domain✅
        9. get_lab_domain✅
        10. assign_class✅
        11. unassign_class✅
        12. generate_timetable
        13. get_domain
        14. select_next_variable
        15. sort_key
        16. backtrack
        17. export_to_excel
        18. print_statistics
    """
    def __init__(self, excel_file='university_database.xlsx'):
        """
        Initialize scheduler by loading all data from Excel file.
        Sets up lookup dictionaries and tracking structures.
        """
        print("\n" + "=" * 80)
        print("LOADING UNIVERSITY DATABASE")
        print("=" * 80)

        # ==================== LOAD ALL DATA FROM EXCEL ====================
        self.rooms = pd.read_excel(excel_file, sheet_name='Rooms').to_dict('records')
        self.labs = pd.read_excel(excel_file, sheet_name='Labs').to_dict('records')
        self.teachers = pd.read_excel(excel_file, sheet_name='Teachers').to_dict('records')
        self.courses = pd.read_excel(excel_file, sheet_name='Courses').to_dict('records')
        self.sections = pd.read_excel(excel_file, sheet_name='Sections').to_dict('records')
        self.assignments = pd.read_excel(excel_file, sheet_name='Course_Assignments').to_dict('records')
        self.time_slots = pd.read_excel(excel_file, sheet_name='Time_Slots').to_dict('records')
        self.days = pd.read_excel(excel_file, sheet_name='Days').to_dict('records')

        # ==================== CREATE FAST LOOKUP DICTIONARIES ====================
        # These allow O(1) lookups instead of scanning through lists
        self.course_dict = {c['Course_ID']: c for c in self.courses}
        self.teacher_dict = {t['Teacher_ID']: t for t in self.teachers}
        self.room_dict = {r['Room_Number']: r for r in self.rooms}
        self.lab_dict = {l['Lab_Name']: l for l in self.labs}
        self.section_dict = {s['Section_ID']: s for s in self.sections}

        # ==================== ORGANIZE LABS BY TYPE ====================
        # Group labs by type for quick filtering (e.g., all Computer labs)
        self.labs_by_type = defaultdict(list)
        for lab in self.labs:
            self.labs_by_type[lab['Lab_Type']].append(lab['Lab_Name'])

        # ==================== TIME SLOT INFORMATION ====================
        self.day_list = [d['Day_ID'] for d in self.days]  # ['MON', 'TUE', ...]
        self.slot_list = [s['Slot_ID'] for s in self.time_slots]  #['S1', 'S2', ...]
        self.slot_times = {s['Slot_ID']: f"{s['Start_Time']} - {s['End_Time']}"
                           for s in self.time_slots}
        # Map slot to index for priority (lower index = earlier/morning slot)
        self.slot_index = {slot_id: idx for idx, slot_id in enumerate(self.slot_list)}

        # ==================== SCHEDULE STORAGE ====================
        self.schedule = []  # Final schedule will be stored here

        # ==================== CONSTRAINT TRACKING STRUCTURES ====================
        # These track what's been scheduled to enforce constraints

        # Track which teachers are busy at each (day, slot)
        self.teacher_schedule = defaultdict(set)  # {(day, slot): {teacher_ids}}

        # Track which rooms/labs are occupied at each (day, slot)
        self.room_schedule = defaultdict(set)  # {(day, slot): {room/lab names}}

        # Track which sections have classes at each (day, slot)
        self.section_schedule = defaultdict(set)  # {(day, slot): {section_ids}}

        # CRITICAL: Track which courses are scheduled on which days per section
        # Prevents same course appearing twice on same day (except consecutive lab slots)
        self.section_course_day = defaultdict(set)  # {(section_id, course_id): {days}}

        # ==================== LOAD BALANCING ====================
        # Track how many classes each section has per day (for even distribution)
        self.section_day_load = defaultdict(lambda: defaultdict(int))  # {section: {day: count}}

        print(f"Loaded: {len(self.rooms)} Rooms, {len(self.labs)} Labs")
        print(f"Loaded: {len(self.teachers)} Teachers, {len(self.courses)} Courses")
        print(f"Loaded: {len(self.sections)} Sections, {len(self.assignments)} Assignments")
        print("=" * 80 + "\n")

    # ==================== HELPER METHODS FOR RESOURCE FINDING ====================

    def get_consecutive_slot_pairs(self):
        """
        Get all pairs of consecutive time slots for lab scheduling.
        Labs require 2 consecutive slots (e.g., S1-S2, S2-S3).
        Returns: List of tuples [(S1, S2), (S2, S3), (S3, S4)]
        """
        pairs = []
        for i in range(len(self.slot_list) - 1):
            pairs.append((self.slot_list[i], self.slot_list[i + 1]))
        return pairs

    def find_suitable_rooms(self, student_count, needs_multimedia):
        """
        Find all rooms that can accommodate the given number of students
        and have multimedia equipment if required.
        """
        suitable = []
        for room in self.rooms:
            # Check capacity
            if room['Strength'] >= student_count:
                # Check multimedia requirement
                if not needs_multimedia or room['Multimedia']:
                    suitable.append(room['Room_Number'])
        return suitable

    def find_suitable_labs(self, lab_type, student_count):
        """
        Find all labs of a specific type that can accommodate students.
        Prioritizes labs with sufficient capacity but includes smaller labs as backup.
        """
        if lab_type not in self.labs_by_type:
            return []

        # Get all labs of this type
        candidate_labs = self.labs_by_type[lab_type]

        # Separate into sufficient and insufficient capacity
        sufficient = []
        insufficient = []
        for lab_name in candidate_labs:
            lab = self.lab_dict[lab_name]
            if lab.get('Strength', 0) >= student_count:
                sufficient.append(lab_name)
            else:
                insufficient.append(lab_name)

        # Return sufficient labs first, then insufficient as backup
        return sufficient + insufficient

    # ==================== CONSTRAINT CHECKING ====================

    def is_slot_available(self, day, slot, section_id, teacher_id, resource):
        """
        Check if a time slot is available for scheduling.
        All three constraints must be satisfied:
        1. Section must be free (students can't be in two places at once)
        2. Teacher must be free (teacher can't teach two classes simultaneously)
        3. Room/Lab must be free (room can't host two classes at once)
        """
        key = (day, slot)

        # Check if section already has a class at this time
        if section_id in self.section_schedule[key]:
            return False

        # Check if teacher is already teaching at this time
        if teacher_id in self.teacher_schedule[key]:
            return False

        # Check if room/lab is already occupied at this time
        if resource in self.room_schedule[key]:
            return False

        return True  # All constraints satisfied

    def is_actual_lab(self, resource_name):
        """
        Check if a resource is an actual lab (not a classroom).
        Used to ensure labs are only scheduled in proper lab spaces.
        """
        return resource_name in self.lab_dict

    def is_course_already_scheduled_today(self, section_id, course_id, day):
        """
        CRITICAL CONSTRAINT: Check if a course is already scheduled on this day.
        Prevents scheduling the same subject multiple times on the same day.
        Exception: Labs can use consecutive slots (2 slots for one lab session).
        """
        return day in self.section_course_day[(section_id, course_id)]

    # ==================== DOMAIN GENERATION ====================
    # Domains = all possible (day, slot, room) combinations for a class

    def get_theory_domain(self, assignment, is_2hour):
        """
            assignment = {
        'Assignment_ID': 101,
        'Section_ID': 'SEC1',
        'Course_ID': 'CSE101',
        'Teacher_ID': 'T1',
        'Student_Count': 40,
        'Type': 'theory'  # or 'lab'    }
        Generate all possible scheduling options for a theory class.
        Returns: List of (day, slot, room) tuples representing valid choices.
        """
        section_id = assignment['Section_ID']
        teacher_id = assignment['Teacher_ID']
        student_count = assignment['Student_Count']
        course = self.course_dict[assignment['Course_ID']]
        needs_multimedia = course['Needs_Multimedia']

        # Find suitable rooms
        suitable_rooms = self.find_suitable_rooms(student_count, needs_multimedia)

        # Determine which time slots are available
        if is_2hour:
            slot_options = ['S4']  # 2-hour classes only fit in S4 slot
        else:
            slot_options = self.slot_list  # Regular classes can use any slot

        # Build all possible (day, slot, room) combinations
        domain = []
        for day in self.day_list:
            for slot in slot_options:
                for room in suitable_rooms:
                    if self.is_slot_available(day, slot, section_id, teacher_id, room):
                        domain.append((day, slot, room))

        return domain

    def get_lab_domain(self, assignment):
        """
        Generate all possible scheduling options for a lab session.
        Labs need 2 consecutive slots.
        Returns: List of (day, slot1, slot2, lab) tuples.
        """
        section_id = assignment['Section_ID']
        teacher_id = assignment['Teacher_ID']
        student_count = assignment['Student_Count']
        course = self.course_dict[assignment['Course_ID']]
        lab_type = course.get('Lab_Type')

        # Find suitable labs
        if not lab_type:
            # No specific type required - any lab with sufficient capacity
            suitable_labs = [l['Lab_Name'] for l in self.labs
                             if l.get('Strength', 0) >= student_count]
        else:
            # Find labs of the required type
            suitable_labs = self.find_suitable_labs(lab_type, student_count)

        # Fallback: If no labs available, use large classrooms
        if not suitable_labs:
            suitable_labs = self.find_suitable_rooms(student_count, needs_multimedia=False)

        # Get consecutive slot pairs
        slot_pairs = self.get_consecutive_slot_pairs()

        # Build all possible (day, slot1, slot2, lab) combinations
        domain = []
        for day in self.day_list:
            for slot1, slot2 in slot_pairs:
                for lab in suitable_labs:
                    # Both consecutive slots must be available
                    if (self.is_slot_available(day, slot1, section_id, teacher_id, lab) and
                            self.is_slot_available(day, slot2, section_id, teacher_id, lab)):
                        domain.append((day, slot1, slot2, lab))

        return domain

    # ==================== SCHEDULE MODIFICATION ====================

    def assign_class(self, day, slot, section_id, teacher_id, resource, assignment_info):
        """
        Assign a class to the schedule and update all tracking structures.
        This ensures constraints are maintained as we build the schedule.
        """
        course_id = assignment_info['Course_ID']

        # Add to the main schedule
        self.schedule.append({
            'Assignment_ID': assignment_info['Assignment_ID'],
            'Section_ID': section_id,
            'Course_ID': course_id,
            'Teacher_ID': teacher_id,
            'Day': day,
            'Slot': slot,
            'Room_or_Lab': resource,
            'Type': assignment_info['Type']
        })

        # Update constraint tracking structures
        key = (day, slot)
        self.section_schedule[key].add(section_id)  # Mark section as busy
        self.teacher_schedule[key].add(teacher_id)  # Mark teacher as busy
        self.room_schedule[key].add(resource)  # Mark room/lab as occupied

        # Mark this course as scheduled on this day for this section
        self.section_course_day[(section_id, course_id)].add(day)

        # Update day load for load balancing
        self.section_day_load[section_id][day] += 1

    def unassign_class(self, day, slot, section_id, teacher_id, resource, course_id):
        """
        Remove a class from the schedule (used during backtracking).
        Reverses all changes made by assign_class.
        """
        # Remove from main schedule
        self.schedule = [c for c in self.schedule
                         if not (c['Day'] == day and c['Slot'] == slot and
                                 c['Section_ID'] == section_id and c['Course_ID'] == course_id)]

        # Update constraint tracking structures
        key = (day, slot)
        self.section_schedule[key].discard(section_id)
        self.teacher_schedule[key].discard(teacher_id)
        self.room_schedule[key].discard(resource)

        # Check if this was the last class of this course on this day
        still_scheduled_today = any(
            c for c in self.schedule
            if c['Section_ID'] == section_id and c['Course_ID'] == course_id and c['Day'] == day
        )
        if not still_scheduled_today:
            self.section_course_day[(section_id, course_id)].discard(day)

        # Update day load
        self.section_day_load[section_id][day] -= 1

    # ==================== CSP BACKTRACKING ALGORITHM ====================

    def generate_timetable(self):
        """
        Main timetable generation using CSP with backtracking.

        Algorithm:
        1. Create a list of all "variables" (classes to schedule)
        2. Use MRV heuristic to select which class to schedule next
        3. Try assigning it to each possible time slot (domain value)
        4. If assignment leads to solution, keep it
        5. If assignment fails, backtrack and try next option
        6. Prioritize morning slots to fill early time slots first
        """
        print("=" * 80)
        print("GENERATING TIMETABLE USING CSP WITH BACKTRACKING")
        print("=" * 80 + "\n")

        # ==================== BUILD VARIABLE LIST ====================
        # Each class session (theory or lab) is a "variable" to assign
        variables = []
        """
            {
                'kind': 'lab',
                'assignment': {
                    'Assignment_ID': 101,
                    'Section_ID': 'SEC-A',
                    'Course_ID': 'CSE101',
                    'Teacher_ID': 'T1',
                    'Student_Count': 40,
                    'Type': 'lab'
                }
            }
        """
        for assignment in self.assignments:
            course = self.course_dict[assignment['Course_ID']]

            # Add one variable for each theory class per week
            for t_idx in range(course['Theory_Classes_Per_Week']):
                variables.append({
                    'kind': 'theory',
                    'assignment': assignment,
                    'index': t_idx
                })

            # Add one variable for lab if course has a lab
            if course.get('Has_Lab'):
                variables.append({
                    'kind': 'lab',
                    'assignment': assignment
                })

        # Shuffle to avoid bias toward certain sections being scheduled first
        random.shuffle(variables)

        # ==================== HELPER FUNCTIONS ====================

        def get_domain(var):
            """Get all possible scheduling options for a variable."""
            assignment = var['assignment']
            if var['kind'] == 'theory':
                course = self.course_dict[assignment['Course_ID']]
                return self.get_theory_domain(assignment, course['Is_2Hour_Special'])
            else:
                return self.get_lab_domain(assignment)

        def select_next_variable(unassigned):
            """
            Select which variable to schedule next using:
            1. Earliest available slot (prioritize morning classes)
            2. MRV (Minimum Remaining Values) as tie-breaker

            This ensures morning slots fill up before afternoon slots.
            """
            best_var = None
            best_domain = None
            best_earliest_slot = 999

            for var in unassigned:
                # Get all possible options for this variable
                domain = get_domain(var)

                # Filter domain by current constraints
                filtered_domain = []

                if var['kind'] == 'theory':
                    # Filter theory class options
                    for (day, slot, room) in domain:
                        # Skip if course already scheduled today
                        if self.is_course_already_scheduled_today(
                                var['assignment']['Section_ID'],
                                var['assignment']['Course_ID'],
                                day
                        ):
                            continue

                        # Skip if slot not available
                        if not self.is_slot_available(
                                day, slot,
                                var['assignment']['Section_ID'],
                                var['assignment']['Teacher_ID'],
                                room
                        ):
                            continue

                        filtered_domain.append((day, slot, room))

                else:  # Lab
                    # Filter lab options (need 2 consecutive slots)
                    for (day, s1, s2, lab) in domain:
                        # Skip if not actual lab
                        if not self.is_actual_lab(lab):
                            continue

                        # Skip if course already scheduled today
                        if self.is_course_already_scheduled_today(
                                var['assignment']['Section_ID'],
                                var['assignment']['Course_ID'],
                                day
                        ):
                            continue

                        # Skip if either slot not available
                        if not (self.is_slot_available(day, s1,
                                                       var['assignment']['Section_ID'],
                                                       var['assignment']['Teacher_ID'], lab) and
                                self.is_slot_available(day, s2,
                                                       var['assignment']['Section_ID'],
                                                       var['assignment']['Teacher_ID'], lab)):
                            continue

                        filtered_domain.append((day, s1, s2, lab))

                # Sort domain: prioritize morning slots, then low-load days
                def sort_key(item):
                    day = item[0]
                    slot = item[1]
                    section = var['assignment']['Section_ID']

                    slot_priority = self.slot_index.get(slot, 999)
                    day_load = self.section_day_load[section][day]

                    return (slot_priority, day_load, random.random())

                filtered_domain = sorted(filtered_domain, key=sort_key)

                # If no options, this variable is most constrained
                if len(filtered_domain) == 0:
                    return var, filtered_domain

                # Find earliest available slot for this variable
                first_option = filtered_domain[0]
                slot = first_option[1]
                earliest_slot = self.slot_index.get(slot, 999)

                # Select variable with earliest available slot
                if best_var is None:
                    best_var = var
                    best_domain = filtered_domain
                    best_earliest_slot = earliest_slot
                elif earliest_slot < best_earliest_slot:
                    # This variable can use earlier slot - prioritize it
                    best_var = var
                    best_domain = filtered_domain
                    best_earliest_slot = earliest_slot
                elif earliest_slot == best_earliest_slot and len(filtered_domain) < len(best_domain):
                    # Same earliest slot, use MRV (smaller domain)
                    best_var = var
                    best_domain = filtered_domain
                    best_earliest_slot = earliest_slot

            return best_var, best_domain

        # ==================== BACKTRACKING SEARCH ====================

        unassigned = variables.copy()

        def backtrack():
            """
            Recursive backtracking function.
            Returns True if all variables successfully assigned, False otherwise.
            """
            # Base case: all variables assigned - success!
            if not unassigned:
                return True

            # Select next variable to assign
            var, domain = select_next_variable(unassigned)

            # No valid options - dead end, backtrack
            if var is None or domain is None or len(domain) == 0:
                return False

            # Remove from unassigned list
            unassigned.remove(var)

            # Try each possible assignment (domain value)
            if var['kind'] == 'theory':
                # Try scheduling theory class
                for day, slot, room in domain:
                    # Prepare assignment info
                    course = self.course_dict[var['assignment']['Course_ID']]
                    assignment_info = {
                        'Assignment_ID': var['assignment']['Assignment_ID'],
                        'Course_ID': var['assignment']['Course_ID'],
                        'Type': '2 Hour Class' if course['Is_2Hour_Special'] else '1.5 Hr Class'
                    }

                    # Make assignment
                    self.assign_class(
                        day, slot,
                        var['assignment']['Section_ID'],
                        var['assignment']['Teacher_ID'],
                        room,
                        assignment_info
                    )

                    # Recursively try to assign remaining variables
                    if backtrack():
                        return True  # Success!

                    # Assignment didn't work, undo it (backtrack)
                    self.unassign_class(
                        day, slot,
                        var['assignment']['Section_ID'],
                        var['assignment']['Teacher_ID'],
                        room,
                        var['assignment']['Course_ID']
                    )

            else:  # Lab
                # Try scheduling lab (2 consecutive slots)
                for day, s1, s2, lab in domain:
                    # Prepare assignment info
                    assignment_info = {
                        'Assignment_ID': var['assignment']['Assignment_ID'],
                        'Course_ID': var['assignment']['Course_ID'],
                        'Type': f"Lab: {lab}"
                    }

                    # Make assignment (both slots)
                    self.assign_class(
                        day, s1,
                        var['assignment']['Section_ID'],
                        var['assignment']['Teacher_ID'],
                        lab,
                        assignment_info
                    )
                    self.assign_class(
                        day, s2,
                        var['assignment']['Section_ID'],
                        var['assignment']['Teacher_ID'],
                        lab,
                        assignment_info
                    )

                    # Recursively try to assign remaining variables
                    if backtrack():
                        return True  # Success!

                    # Assignment didn't work, undo it (backtrack)
                    self.unassign_class(
                        day, s1,
                        var['assignment']['Section_ID'],
                        var['assignment']['Teacher_ID'],
                        lab,
                        var['assignment']['Course_ID']
                    )
                    self.unassign_class(
                        day, s2,
                        var['assignment']['Section_ID'],
                        var['assignment']['Teacher_ID'],
                        lab,
                        var['assignment']['Course_ID']
                    )

            # No assignment worked, restore variable and return failure
            unassigned.append(var)
            return False

        # ==================== RUN BACKTRACKING ====================
        success = backtrack()

        # ==================== CALCULATE RESULTS ====================
        total = len(self.assignments)
        failed = []
        success_count = 0

        for assignment in self.assignments:
            course = self.course_dict[assignment['Course_ID']]
            theory_needed = course['Theory_Classes_Per_Week']

            # Count scheduled theory classes
            theory_count = len([c for c in self.schedule
                                if c['Assignment_ID'] == assignment['Assignment_ID']
                                and 'Lab' not in c['Type']])

            # Check if lab scheduled (if required)
            lab_ok = True
            if course.get('Has_Lab'):
                lab_ok = any(c for c in self.schedule
                             if c['Assignment_ID'] == assignment['Assignment_ID']
                             and 'Lab' in c['Type'])

            # Mark as success or failure
            if theory_count >= theory_needed and lab_ok:
                success_count += 1
            else:
                failed.append(assignment)

        # ==================== PRINT SUMMARY ====================
        print("=" * 80)
        print("GENERATION SUMMARY")
        print("=" * 80)
        print(f"Total Assignments: {total}")
        print(f"Successfully Scheduled: {success_count}")
        print(f"Failed: {len(failed)}")
        print(f"Success Rate: {success_count / total * 100:.1f}%")
        print(f"Total Classes Scheduled: {len(self.schedule)}")
        print("=" * 80 + "\n")

        return success_count, failed

    # ==================== EXPORT TO EXCEL ====================

    def export_to_excel(self):
        """Export generated timetables to Excel file with nice formatting."""
        print("=" * 80)
        print("EXPORTING TIMETABLES TO EXCEL")
        print("=" * 80 + "\n")

        wb = Workbook()
        wb.remove(wb.active)

        # Day name mapping
        day_names = {
            'MON': 'Monday',
            'TUE': 'Tuesday',
            'WED': 'Wednesday',
            'THU': 'Thursday',
            'FRI': 'Friday'
        }

        # Create one sheet per section
        print("Creating section timetables...")
        for section in self.sections:
            section_id = section['Section_ID']
            ws = wb.create_sheet(title=f"Section_{section_id}")

            # ==================== CREATE HEADER ROW ====================
            headers = ['Day\\Time'] + [self.slot_times[s] for s in self.slot_list]
            ws.append(headers)

            # Style header
            for cell in ws[1]:
                cell.font = Font(bold=True, size=11, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092",
                                        fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center',
                                           wrap_text=True)
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )

            # Set column widths
            ws.column_dimensions['A'].width = 15  # Day column
            for i in range(2, len(headers) + 1):
                ws.column_dimensions[chr(64 + i)].width = 28  # Time slot columns

            # ==================== FILL TIMETABLE DATA ====================
            for day_id in self.day_list:
                row_data = [day_names[day_id]]

                for slot_id in self.slot_list:
                    # Find all classes for this section at this day/time
                    classes = [c for c in self.schedule
                               if c['Section_ID'] == section_id
                               and c['Day'] == day_id
                               and c['Slot'] == slot_id]

                    if classes:
                        # Format: Course Name\nRoom\nTeacher
                        parts = []
                        for cls in classes:
                            course = self.course_dict[cls['Course_ID']]
                            teacher = self.teacher_dict[cls['Teacher_ID']]
                            part = (f"{course['Course_Name']}\n"
                                    f"{cls['Room_or_Lab']}\n"
                                    f"{teacher['Teacher_Name']}")
                            parts.append(part)
                        cell_content = "\n---\n".join(parts)
                    else:
                        cell_content = ""  # Empty slot

                    row_data.append(cell_content)

                ws.append(row_data)

                # Style data row
                current_row = ws.max_row
                ws.row_dimensions[current_row].height = 70  # Tall rows for readability

                for cell in ws[current_row]:
                    cell.alignment = Alignment(horizontal='left', vertical='top',
                                               wrap_text=True)
                    cell.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )
                    # Highlight day column
                    if cell.column == 1:
                        cell.font = Font(bold=True, size=11)
                        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6",
                                                fill_type="solid")

        # Create teacher timetables
        print("Creating teacher timetables...")
        for teacher in self.teachers:
            teacher_id = teacher['Teacher_ID']
            ws = wb.create_sheet(title=f"Teacher_{teacher_id}")

            # Header row
            headers = ['Day\\Time'] + [self.slot_times[s] for s in self.slot_list]
            ws.append(headers)

            # Style header
            for cell in ws[1]:
                cell.font = Font(bold=True, size=11, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092",
                                        fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center',
                                           wrap_text=True)
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )

            # Set column widths
            ws.column_dimensions['A'].width = 15  # Day column
            for i in range(2, len(headers) + 1):
                ws.column_dimensions[chr(64 + i)].width = 28  # Time slot columns

            # Fill timetable data for this teacher
            for day_id in self.day_list:
                row_data = [day_names[day_id]]

                for slot_id in self.slot_list:
                    # Find all classes for this teacher at this day/time
                    classes = [c for c in self.schedule
                               if c['Teacher_ID'] == teacher_id
                               and c['Day'] == day_id
                               and c['Slot'] == slot_id]

                    if classes:
                        # Format: Course Name\nSection\nRoom
                        parts = []
                        for cls in classes:
                            course = self.course_dict[cls['Course_ID']]
                            part = (f"{course['Course_Name']}\n"
                                    f"Section: {cls['Section_ID']}\n"
                                    f"{cls['Room_or_Lab']}")
                            parts.append(part)
                        cell_content = "\n---\n".join(parts)
                    else:
                        cell_content = ""  # Empty slot

                    row_data.append(cell_content)

                ws.append(row_data)

                # Style data row
                current_row = ws.max_row
                ws.row_dimensions[current_row].height = 70  # Tall rows for readability

                for cell in ws[current_row]:
                    cell.alignment = Alignment(horizontal='left', vertical='top',
                                               wrap_text=True)
                    cell.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )
                    # Highlight day column
                    if cell.column == 1:
                        cell.font = Font(bold=True, size=11)
                        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6",
                                                fill_type="solid")

        # Save Excel file
        wb.save('generated_timetable_csp.xlsx')
        print("✓ Timetables exported: generated_timetable_csp.xlsx")
        print("=" * 80 + "\n")

    # ==================== STATISTICS ====================

    def print_statistics(self):
        """Print detailed statistics about the generated timetable."""
        print("=" * 80)
        print("TIMETABLE STATISTICS")
        print("=" * 80 + "\n")

        print("SECTION LOAD DISTRIBUTION:")
        print("-" * 80)
        for section in self.sections:
            section_id = section['Section_ID']
            day_loads = self.section_day_load[section_id]
            total = sum(day_loads.values())

            # Show classes per day
            loads_str = ", ".join([f"{d}:{day_loads[d]}" for d in self.day_list])
            print(f"  {section_id}: {loads_str} | Total: {total}")

        print("\n" + "-" * 80)

        # Count different types of classes
        lab_slots = len([c for c in self.schedule if 'Lab' in c['Type']])
        theory_classes = len([c for c in self.schedule if 'Lab' not in c['Type']])

        print(f"Lab Sessions: {lab_slots // 2} (using {lab_slots} slots)")
        print(f"Theory Classes: {theory_classes}")
        print(f"Total Scheduled Slots: {len(self.schedule)}")

        print("\n" + "=" * 80 + "\n")

